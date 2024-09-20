import os
from openpyxl import load_workbook
from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog, QTableWidget, QHeaderView, QTableWidgetItem
from PyQt6.QtGui import QFont
from db.conexion import ConexionMysql
from db.querys import Query

class Menu(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = ConexionMysql()
        self.menu = uic.loadUi('views/menu.ui')
        self.menu.show()
        self.menu.btnImoto.clicked.connect(self.insertMoto)
        self.menu.btnInsert.clicked.connect(self.insertRepuesto)
        self.menu.btnImagen.clicked.connect(self.openImagen)
        self.menu.btnIlistado.clicked.connect(self.openExcel)
        self.menu.btnInsertlist.clicked.connect(self.insertList)
        self.category = []
        self.moto = []
        self.img = []
        self.showTable()
        self.showTlistado()
        self.showTableRepuestos()
        self.showCategory()
        self.showClistado()
        self.showMoto()
        self.showMlistado()
        
        
    def showTable(self):
        columns = ['MOTO', 'DESCRIPCION', 'MODELO', 'MARCA']
        self.menu.tMoto.setFont(QFont("FiraCode Nerd Font", 12))
        self.menu.tMoto.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.menu.tMoto.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.menu.tMoto.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.menu.tMoto.setStyleSheet(header_style)
    
    def showTlistado(self):
        columns = ['MOTO', 'DESCRIPCION', 'CATEGORIA', 'MOTO']
        self.menu.tRlistado.setFont(QFont("FiraCode Nerd Font", 12))
        self.menu.tRlistado.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.menu.tRlistado.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.menu.tRlistado.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.menu.tRlistado.setStyleSheet(header_style)
    
    def showTableRepuestos(self):
        columns = ['CÓDIGO', 'DESCRIPCION', 'CATEGORIA', 'MOTO']
        self.menu.tRepuestos.setFont(QFont("FiraCode Nerd Font", 12))
        self.menu.tRepuestos.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.menu.tRepuestos.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.menu.tRepuestos.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.menu.tRepuestos.setStyleSheet(header_style)

    
    def showCategory(self):
        query = Query()
        options = set(self.category)
        result = query.selectCategory()
        for data, datos in enumerate(result):
            options.add(str(datos[1]))
        self.menu.cbCategoria.addItems(list(options))
    
    def showClistado(self):
        query = Query()
        options = set(self.category)
        result = query.selectCategory()
        for id, data in result:
            self.menu.cbClistado.addItem(str(data))
        
    
    def showMoto(self):
        query = Query()
        result = query.selectMoto()
        options = set(self.moto)
        print(result)
        for id, data in result:
            print(id)
            print(data)
            self.menu.cbMlistado.addItem(str(data))
        """for data, datos in enumerate(result):
            print(datos)
            options.add(str(datos[1]))
            
        selects = list(options)
        selects.sort()
        self.menu.cbMoto.addItems(selects)"""
        
    def showMlistado(self):
        query = Query()
        result = query.selectMoto()
        options = set(self.moto)
        print(result)
        for id, data in result:
            print(id)
            print(data)
            self.menu.cbMoto.addItem(str(data))
            
    def insertMoto(self):
        self.descripcion = self.menu.txtDmoto.toPlainText().strip()
        self.nombre = self.menu.txtName.text()
        self.modelo = int(self.menu.cbYear.currentText())
        self.marca = self.menu.txtMarca.text()
        query = Query()
        if self.descripcion:
            query.insertMoto(self.nombre, self.descripcion, self.modelo, self.marca)
        else: 
            self.descripcion = ''
            query.insertMoto(self.nombre, self.descripcion, self.modelo, self.marca)
            
        row = self.menu.tMoto.rowCount()
        self.menu.tMoto.insertRow(row)
        self.menu.tMoto.setItem(row, 0, QTableWidgetItem(self.nombre))
        self.menu.tMoto.setItem(row, 1, QTableWidgetItem(self.descripcion))
        self.menu.tMoto.setItem(row, 2, QTableWidgetItem(str(self.modelo)))
        self.menu.tMoto.setItem(row, 3, QTableWidgetItem(self.marca))
        self.menu.cbMlistado.clear()
        self.showMoto()
        self.menu.cbMoto.clear()
        self.menu.txtName.setText("")
        self.menu.cbYear.clear()
        self.menu.txtMarca.setText("")
        self.menu.txtDmoto.setText("")
            
    def insertRepuesto(self):
        query = Query()
        idcategory = self.menu.cbCategoria.currentIndex()+1
        idmoto = self.menu.cbMoto.currentIndex()+1
        if self.img:
            for ruta_imagen in self.img:  
                with open(ruta_imagen, 'rb') as file:
                    print(ruta_imagen)
                    imagen_binaria = file.read()
                    query.insertRepuesto(self.menu.txtCodigo.text(), self.menu.txtDescription.toPlainText(), imagen_binaria, idcategory, idmoto)
        else:
            query.insertRepuesto(self.menu.txtCodigo.text(), self.menu.txtDescription.toPlainText(), None, idcategory, idmoto)
            
        self.db.close_connection()
        row = self.menu.tRepuestos.rowCount()
        self.menu.tRepuestos.insertRow(row)
        self.menu.tRepuestos.setItem(row, 0, QTableWidgetItem(self.menu.txtCodigo.text()))
        self.menu.tRepuestos.setItem(row, 1, QTableWidgetItem(self.menu.txtDescription.toPlainText()))
        self.menu.tRepuestos.setItem(row, 2, QTableWidgetItem(self.menu.cbCategoria.currentText()))
        self.menu.tRepuestos.setItem(row, 3, QTableWidgetItem(self.menu.cbMoto.currentText()))
    
    def loadList(self, path):
        workbook = load_workbook(filename=path)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only= True):
            row_index = self.menu.tRlistado.rowCount()
            cod = row[0].replace('-','')
            self.menu.tRlistado.insertRow(row_index)
            print(row[1])
            self.menu.tRlistado.setItem(row_index, 0, QTableWidgetItem(str(cod)))
            self.menu.tRlistado.setItem(row_index, 1, QTableWidgetItem(str(row[1])))
            self.menu.tRlistado.setItem(row_index, 2, QTableWidgetItem(str(self.menu.cbClistado.currentIndex()+1)))
            self.menu.tRlistado.setItem(row_index, 3, QTableWidgetItem(str(self.menu.cbMlistado.currentIndex()+1)))
            
    def insertList(self):
        query = Query()
        try: 
            for row in range(self.menu.tRlistado.rowCount()):
                cod = self.menu.tRlistado.item(row, 0)
                description = self.menu.tRlistado.item(row, 1)
                category = self.menu.tRlistado.item(row, 2)
                moto = self.menu.tRlistado.item(row, 3)
                query.insertRepuesto(cod.text(),description.text(), None, int(category.text()), int(moto.text()))
                self.menu.cbMlistado.clear()
        except Exception as e:
            print(e)
        finally:
            self.db.close_connection()
    
    def openImagen(self):
        folder = QFileDialog()
        folder_path, __= folder.getOpenFileNames(None, 'Cargar imagen', '', 'JPEG (*.jpg)')
        self.img = folder_path
    
    def openExcel(self):
        folder = QFileDialog()
        folder_path, __= folder.getOpenFileName(None, 'ABRIR ARCHIVO', '', 'xlsx (*.xlsx)')
        print(folder_path)
        self.loadList(folder_path)
        